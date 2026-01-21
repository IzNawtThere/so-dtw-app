"""
NeoAsia Sales Order DTW Web Application
========================================
A Streamlit web app for generating SAP DTW import files.

Upload an Excel file with Sales Orders and Line Items,
validate the data, and download properly formatted DTW files.

Version: 1.0.0
"""

import streamlit as st
import pandas as pd
import io
import zipfile
from datetime import datetime
from typing import Tuple, List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# PAGE CONFIGURATION
# ============================================================================
st.set_page_config(
    page_title="NeoAsia Sales Order DTW Tool",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================================================
# CONSTANTS - DTW FILE STRUCTURE
# ============================================================================
ORDR_TOTAL_COLS = 271
RDR1_TOTAL_COLS = 244

# ORDR field positions (0-based)
ORDR_FIELDS = {
    'DocNum': 0,
    'DocType': 2,
    'DocDate': 5,
    'DocDueDate': 6,
    'CardCode': 7,
    'SalesPersonCode': 21,
    'U_AllowDOInv_Dep': 198,
    'U_GSTSGDRate': 202,
}

# RDR1 field positions (0-based)
RDR1_FIELDS = {
    'ParentKey': 0,
    'LineNum': 1,
    'ItemCode': 2,
    'Quantity': 4,
    'Price': 6,
    'WarehouseCode': 13,
    'SalesPersonCode': 14,
    'AccountCode': 17,
    'CostingCode': 20,
    'VatGroup': 23,
    'COGSCostingCode': 81,
    'CostingCode2': 87,
    'CostingCode3': 88,
    'CostingCode4': 89,
    'CostingCode5': 90,
    'COGSCostingCode2': 97,
    'COGSCostingCode3': 98,
    'COGSCostingCode4': 99,
    'COGSCostingCode5': 100,
    'U_PermitNum': 171,
    'U_GSTSGDRate': 183,
}

# ============================================================================
# HEADER DEFINITIONS (Complete DTW structure)
# ============================================================================
ORDR_HEADER_ROW1 = [
    "DocNum", "DocEntry", "DocType", "HandWritten", "Printed", "DocDate", "DocDueDate", "CardCode", "CardName", "Address",
    "NumAtCard", "DocTotal", "AttachmentEntry", "DocCurrency", "DocRate", "Reference1", "Reference2", "Comments", "JournalMemo", "PaymentGroupCode",
    "DocTime", "SalesPersonCode", "TransportationCode", "Confirmed", "ImportFileNum", "SummeryType", "ContactPersonCode", "ShowSCN", "Series", "TaxDate",
    "PartialSupply", "DocObjectCode", "ShipToCode", "Indicator", "FederalTaxID", "DiscountPercent", "PaymentReference", "DocTotalFc", "Form1099", "Box1099",
    "RevisionPo", "RequriedDate", "CancelDate", "BlockDunning", "Pick", "PaymentMethod", "PaymentBlock", "PaymentBlockEntry", "CentralBankIndicator", "MaximumCashDiscount",
    "Project", "ExemptionValidityDateFrom", "ExemptionValidityDateTo", "WareHouseUpdateType", "Rounding", "ExternalCorrectedDocNum", "InternalCorrectedDocNum", "DeferredTax", "TaxExemptionLetterNum", "AgentCode",
    "NumberOfInstallments", "ApplyTaxOnFirstInstallment", "VatDate", "DocumentsOwner", "FolioPrefixString", "FolioNumber", "DocumentSubType", "BPChannelCode", "BPChannelContact", "Address2",
    "PayToCode", "ManualNumber", "UseShpdGoodsAct", "IsPayToBank", "PayToBankCountry", "PayToBankCode", "PayToBankAccountNo", "PayToBankBranch", "BPL_IDAssignedToInvoice", "DownPayment",
    "ReserveInvoice", "LanguageCode", "TrackingNumber", "PickRemark", "ClosingDate", "SequenceCode", "SequenceSerial", "SeriesString", "SubSeriesString", "SequenceModel",
    "UseCorrectionVATGroup", "DownPaymentAmount", "DownPaymentPercentage", "DownPaymentType", "DownPaymentAmountSC", "DownPaymentAmountFC", "VatPercent", "ServiceGrossProfitPercent", "OpeningRemarks", "ClosingRemarks",
    "RoundingDiffAmount", "ControlAccount", "InsuranceOperation347", "ArchiveNonremovableSalesQuotation", "GTSChecker", "GTSPayee", "ExtraMonth", "ExtraDays", "CashDiscountDateOffset", "StartFrom",
    "NTSApproved", "ETaxWebSite", "ETaxNumber", "NTSApprovedNumber", "EDocGenerationType", "EDocSeries", "EDocNum", "EDocExportFormat", "EDocStatus", "EDocErrorCode",
    "EDocErrorMessage", "DownPaymentStatus", "GroupSeries", "GroupNumber", "GroupHandWritten", "ReopenOriginalDocument", "ReopenManuallyClosedOrCanceledDocument", "CreateOnlineQuotation", "POSEquipmentNumber", "POSManufacturerSerialNumber",
    "POSCashierNumber", "ApplyCurrentVATRatesForDownPaymentsToDraw", "ClosingOption", "SpecifiedClosingDate", "OpenForLandedCosts", "RelevantToGTS", "AnnualInvoiceDeclarationReference", "Supplier", "Releaser", "Receiver",
    "BlanketAgreementNumber", "IsAlteration", "AssetValueDate", "DocumentDelivery", "AuthorizationCode", "StartDeliveryDate", "StartDeliveryTime", "EndDeliveryDate", "EndDeliveryTime", "VehiclePlate",
    "ATDocumentType", "ElecCommStatus", "ReuseDocumentNum", "ReuseNotaFiscalNum", "PrintSEPADirect", "FiscalDocNum", "POSDailySummaryNo", "POSReceiptNo", "PointOfIssueCode", "Letter",
    "FolioNumberFrom", "FolioNumberTo", "InterimType", "RelatedType", "RelatedEntry", "SAPPassport", "DocumentTaxID", "DateOfReportingControlStatementVAT", "ReportingSectionControlStatementVAT", "ExcludeFromTaxReportControlStatementVAT",
    "POS_CashRegister", "CreateQRCodeFrom", "PriceMode", "CommissionTrade", "CommissionTradeReturn", "UseBillToAddrToDetermineTax", "Cig", "Cup", "FatherCard", "FatherType",
    "ShipState", "ShipPlace", "CustOffice", "FCI", "AddLegIn", "LegTextF", "IndFinal", "U_LoanExpDate", "U_Transfer_Type", "U_IGN_RsvDue",
    "U_IGN_RsvSONum", "U_PCH_ApprvPay", "U_RDR_FinancingAppr", "U_Star_Acct", "U_AllowDOInv_Dep", "U_CustPO_Received", "U_RCN_Num", "U_POR_InventoryValue", "U_Appr_LowSlsPrice", "U_Appr_PO_Stock",
    "U_Appr_PO_NStock1", "U_Appr_PO_NStock2", "U_GSTSGDRate", "U_GSTSGDAmt", "U_Vessel", "U_LineBusiness", "U_RDR_IgnoreSpcTerms", "U_Dep_Required", "U_Dep_Received", "U_LoanAddr",
    "U_TRF_SlpName", "U_INM_Purpose", "U_ELTO_RevSchCrt", "U_ELTO_RevSchVal", "U_ELTO_RevCrtdDat", "U_ELTO_BillSchCrt", "U_ELTO_BillSchVal", "U_ELTO_BillCrtDat", "U_ELTO_LseTrf", "U_ELTO_LseTransNo",
    "U_Appr_DiffSQ", "U_FIN_TotalBefDisc", "U_FIN_LessDisc", "U_FIN_AddGST", "U_FIN_LessDeposit", "U_FIN_LessDepGST", "U_RefreshInfo", "U_PL_Remarks", "U_RDR_SellPrcSample", "U_Appr_Sample",
    "U_FIN_GSTDueFmFinc", "U_IntRemarks", "U_ImportConsign", "U_EndUserCode", "U_EndUserName", "U_TRF_AccList", "U_ELTO_RSVCode", "U_ELTO_ContactDetail", "U_ELTO_RsvComment", "U_FileName",
    "U_RDocNum", "U_Warehouse_Code", "U_Warehouse", "U_P_Approval_Code", "U_Customer_State", "U_Cust_Territory", "U_Team_Code", "U_Detailman_Code", "U_ContractNo", "U_ApplNeoComm",
    "U_SampleInv", "U_ETA", "U_RecEntry", "U_WMS_BasePick", "U_AutoPick", "U_URGENT", "U_WMS_RefNo", "U_WMS_FromHandheld", "U_WMS_DeviceID", "U_WMS_ScnUsr",
    "U_DeliveryCompt", "U_DeliveryDate", "U_CAPPSUpdate", "U_CAPPSUpdDat", "U_CAPPSNum", "U_OrigFile", "U_OrigName", "U_ServFile", "U_ServName", "U_CAPPSStsUpd",
    "U_CAPPSStsDat"
]

ORDR_HEADER_ROW2 = [
    "DocNum", "DocEntry", "DocType", "Handwrtten", "Printed", "DocDate", "DocDueDate", "CardCode", "CardName", "Address",
    "NumAtCard", "DocTotal", "AtcEntry", "DocCur", "DocRate", "Ref1", "Ref2", "Comments", "JrnlMemo", "GroupNum",
    "DocTime", "SlpCode", "TrnspCode", "Confirmed", "ImportEnt", "SummryType", "CntctCode", "ShowSCN", "Series", "TaxDate",
    "PartSupply", "ObjType", "ShipToCode", "Indicator", "LicTradNum", "DiscPrcnt", "PaymentRef", "UserSign", "DocTotalFC", "Form1099",
    "Box1099", "RevisionPo", "ReqDate", "CancelDate", "BlockDunn", "Pick", "PeyMethod", "PayBlock", "PayBlckRef", "CntrlBnk",
    "MaxDscn", "Project", "FromDate", "ToDate", "UpdInvnt", "Rounding", "CorrExt", "CorrInv", "DeferrTax", "LetterNum",
    "AgentCode", "Installmnt", "VATFirst", "VatDate", "OwnerCode", "FolioPref", "FolioNum", "DocSubType", "BPChCode", "BPChCntc",
    "Address2", "PayToCode", "ManualNum", "UseShpdGd", "IsPaytoBnk", "BnkCntry", "BankCode", "BnkAccount", "BnkBranch", "BPLId",
    "DpmPrcnt", "isIns", "LangCode", "TrackNo", "PickRmrk", "ClsDate", "SeqCode", "Serial", "SeriesStr", "SubStr",
    "Model", "UseCorrVat", "DpmAmnt", "DpmPrcnt", "Posted", "DpmAmntSC", "DpmAmntFC", "VatPercent", "SrvGpPrcnt", "Header",
    "Footer", "RoundDif", "CtlAccount", "InsurOp347", "IgnRelDoc", "Checker", "Payee", "ExtraMonth", "ExtraDays", "CdcOffset",
    "PayDuMonth", "NTSApprov", "NTSWebSite", "NTSeTaxNo", "NTSApprNo", "EDocGenTyp", "ESeries", "EDocNum", "EDocExpFrm", "EDocStatus",
    "EDocErrCod", "EDocErrMsg", "DpmStatus", "PQTGrpSer", "PQTGrpNum", "PQTGrpHW", "ReopOriDoc", "ReopManCls", "OnlineQuo", "POSEqNum",
    "POSManufSN", "POSCashN", "DpmAsDscnt", "ClosingOpt", "SpecDate", "OpenForLaC", "GTSRlvnt", "AnnInvDecR", "Supplier", "Releaser",
    "Receiver", "AgrNo", "IsAlt", "AssetDate", "DocDlvry", "AuthCode", "StDlvDate", "StDlvTime", "EndDlvDate", "EndDlvTime",
    "VclPlate", "AtDocType", "ElCoStatus", "IsReuseNum", "IsReuseNFN", "PrintSEPA", "FiscDocNum", "ZrdAbs", "POSRcptNo", "PTICode",
    "Letter", "FolNumFrom", "FolNumTo", "InterimTyp", "RelatedTyp", "RelatedEnt", "SAPPassprt", "DocTaxID", "DateReport", "RepSection",
    "ExclTaxRep", "PosCashReg", "QRCodeSrc", "PriceMode", "ShipToCode", "ComTrade", "ComTradeRt", "UseBilAddr", "CIG", "CUP",
    "FatherCard", "FatherType", "ShipState", "ShipPlace", "CustOffice", "FCI", "AddLegIn", "LegTextF", "DANFELgTxt", "IndFinal",
    "DataVers", "U_LoanExpDate", "U_Transfer_Type", "U_IGN_RsvDue", "U_IGN_RsvSONum", "U_PCH_ApprvPay", "U_RDR_FinancingAppr", "U_Star_Acct", "U_AllowDOInv_Dep", "U_CustPO_Received",
    "U_RCN_Num", "U_POR_InventoryValue", "U_Appr_LowSlsPrice", "U_Appr_PO_Stock", "U_Appr_PO_NStock1", "U_Appr_PO_NStock2", "U_GSTSGDRate", "U_GSTSGDAmt", "U_Vessel", "U_LineBusiness",
    "U_RDR_IgnoreSpcTerms", "U_Dep_Required", "U_Dep_Received", "U_LoanAddr", "U_TRF_SlpName", "U_INM_Purpose", "U_ELTO_RevSchCrt", "U_ELTO_RevSchVal", "U_ELTO_RevCrtdDat", "U_ELTO_BillSchCrt",
    "U_ELTO_BillSchVal", "U_ELTO_BillCrtDat", "U_ELTO_LseTrf", "U_ELTO_LseTransNo", "U_Appr_DiffSQ", "U_FIN_TotalBefDisc", "U_FIN_LessDisc", "U_FIN_AddGST", "U_FIN_LessDeposit", "U_FIN_LessDepGST",
    "U_RefreshInfo", "U_PL_Remarks", "U_RDR_SellPrcSample", "U_Appr_Sample", "U_FIN_GSTDueFmFinc", "U_IntRemarks", "U_ImportConsign", "U_EndUserCode", "U_EndUserName", "U_TRF_AccList",
    "U_ELTO_RSVCode", "U_ELTO_ContactDetail", "U_ELTO_RsvComment", "U_FileName", "U_RDocNum", "U_Warehouse_Code", "U_Warehouse", "U_P_Approval_Code", "U_Customer_State", "U_Cust_Territory",
    "U_Team_Code", "U_Detailman_Code", "U_ContractNo", "U_ApplNeoComm", "U_SampleInv", "U_ETA", "U_RecEntry", "U_WMS_BasePick", "U_AutoPick", "U_URGENT",
    "U_WMS_RefNo", "U_WMS_FromHandheld", "U_WMS_DeviceID", "U_WMS_ScnUsr", "U_DeliveryCompt", "U_DeliveryDate", "U_CAPPSUpdate", "U_CAPPSUpdDat", "U_CAPPSNum", "U_OrigFile",
    "U_OrigName"
]

RDR1_HEADER_ROW1 = [
    "ParentKey", "LineNum", "ItemCode", "ItemDescription", "Quantity", "ShipDate", "Price", "PriceAfterVAT", "Currency", "Rate",
    "DiscountPercent", "VendorNum", "SerialNum", "WarehouseCode", "SalesPersonCode", "CommisionPercent", "TreeType", "AccountCode", "UseBaseUnits", "SupplierCatNum",
    "CostingCode", "ProjectCode", "BarCode", "VatGroup", "Height1", "Hight1Unit", "Height2", "Height2Unit", "Lengh1", "Lengh1Unit",
    "Lengh2", "Lengh2Unit", "Weight1", "Weight1Unit", "Weight2", "Weight2Unit", "Factor1", "Factor2", "Factor3", "Factor4",
    "BaseType", "BaseEntry", "BaseLine", "Volume", "VolumeUnit", "Width1", "Width1Unit", "Width2", "Width2Unit", "Address",
    "TaxCode", "TaxType", "TaxLiable", "BackOrder", "FreeText", "ShippingMethod", "CorrectionInvoiceItem", "CorrInvAmountToStock", "CorrInvAmountToDiffAcct", "WTLiable",
    "DeferredTax", "MeasureUnit", "UnitsOfMeasurment", "LineTotal", "TaxPercentagePerRow", "TaxTotal", "ConsumerSalesForecast", "ExciseAmount", "CountryOrg", "SWW",
    "TransactionType", "DistributeExpense", "RowTotalFC", "CFOPCode", "CSTCode", "Usage", "TaxOnly", "UnitPrice", "LineStatus", "PackageQuantity",
    "LineType", "COGSCostingCode", "COGSAccountCode", "ChangeAssemlyBoMWarehouse", "GrossBuyPrice", "GrossBase", "GrossProfitTotalBasePrice", "CostingCode2", "CostingCode3", "CostingCode4",
    "CostingCode5", "ItemDetails", "LocationCode", "ActualDeliveryDate", "ExLineNo", "RequiredDate", "RequiredQuantity", "COGSCostingCode2", "COGSCostingCode3", "COGSCostingCode4",
    "COGSCostingCode5", "CSTforIPI", "CSTforPIS", "CSTforCOFINS", "CreditOriginCode", "WithoutInventoryMovement", "AgreementNo", "AgreementRowNumber", "ActualBaseEntry", "ActualBaseLine",
    "DocEntry", "Surpluses", "DefectAndBreakup", "Shortages", "ConsiderQuantity", "PartialRetirement", "RetirementQuantity", "RetirementAPC", "ThirdParty", "PoNum",
    "PoItmNum", "ExpenseType", "ReceiptNumber", "ExpenseOperationType", "FederalTaxID", "GrossProfit", "GrossProfitFC", "GrossProfitSC", "UoMEntry", "InventoryQuantity",
    "ParentLineNum", "Incoterms", "TransportMode", "NatureOfTransaction", "DestinationCountryForImport", "DestinationRegionForImport", "OriginCountryForExport", "OriginRegionForExport", "ChangeInventoryQuantityIndependently", "FreeOfChargeBP",
    "SACEntry", "HSNEntry", "GrossPrice", "GrossTotal", "GrossTotalFC", "NCMCode", "NVECode", "IndEscala", "CtrSealQty", "CNJPMan",
    "CESTCode", "UFFiscalBenefitCode", "ReverseCharge", "ShipToCode", "ShipToDescription", "OwnerCode", "ExternalCalcTaxRate", "ExternalCalcTaxAmount", "StandardItemIdentification", "CommodityClassification",
    "UnencumberedReason", "CUSplit", "U_ZL_SlpCode", "U_ZL_SlpName", "U_ZL_CardCode", "U_ZL_CardName", "U_ZL_InvoiceNum", "U_ZL_ItemCode", "U_ZL_ItemName", "U_BonusItem",
    "U_HasAlternativeItem", "U_PermitNum", "U_Has_EPoint", "U_PI_Qty", "U_PI_Price", "U_Redemption", "U_EPoint_Redeemed", "U_Rebate_RefNum", "U_SLS_Remarks1", "U_SLS_Remarks2",
    "U_SLS_Remarks3", "U_SLS_Remarks4", "U_SLS_Remarks5", "U_GSTSGDRate", "U_GSTSGDAmt", "U_Related_SO_Number", "U_Related_SO_LineNum", "U_QUT_Calc_BOM_Cost", "U_RDR_HasSpcTerms", "U_RDR_SpcPayTerms",
    "U_TRF_SerialNum", "U_ELTO_DefIncInd", "U_ELTO_ConStrDat", "U_ELTO_ConEndDat", "U_ELTO_FreqBill", "U_ELTO_FreqAdvRev", "U_ELTO_BillStrDat", "U_ELTO_LseQty", "U_ELTO_CustPO", "U_ELTO_EstBillAmt",
    "U_ELTO_BillId", "U_ELTO_BillLine", "U_ELTO_NoCustPOApp", "U_ELTO_Lease", "U_ELTO_LseItmCde", "U_ELTO_LseItmDsc", "U_ELTO_LseWhs", "U_ELTO_LseDlvQty", "U_ELTO_LseRtnQty", "U_ELTO_RevSchType",
    "U_ELTO_ContractQty", "U_ELTO_QtyConPrice", "U_ELTO_Prorata", "U_ELTO_BillPeriodFr", "U_ELTO_BillPeriodTo", "U_ELTO_BaseDocEntry", "U_ELTO_BaseDocNum", "U_ELTO_BaseRow", "U_FIN_Shown", "U_PL_CartonNum",
    "U_PL_KgCarton", "U_PL_MYRegNum", "U_TRF_VIT_Prc", "U_TRF_SellingPrice", "U_UnitPriceBP", "U_POCountry", "U_Invoice_No", "U_Invoice_Item", "U_Return_Ref_No", "U_P_Customer_Code",
    "U_ZP_Item_Code", "U_Credit_Reason", "U_List_Price", "U_ZP_Customer_Code", "U_Customer_Name", "U_ZP_Invoice_Item", "U_Customer_State", "U_Cust_Territory", "U_TempDivision", "U_RecLines",
    "U_COO", "U_TfrBinTo", "U_TrfFromBin", "U_GRN"
]

RDR1_HEADER_ROW2 = [
    "DocNum", "LineNum", "ItemCode", "Dscription", "Quantity", "ShipDate", "Price", "PriceAfVAT", "Currency", "Rate",
    "DiscPrcnt", "VendorNum", "SerialNum", "WhsCode", "SlpCode", "Commission", "TreeType", "AcctCode", "UseBaseUn", "SubCatNum",
    "OcrCode", "Project", "CodeBars", "VatGroup", "Height1", "Hght1Unit", "Height2", "Hght2Unit", "Length1", "Len1Unit",
    "length2", "Len2Unit", "Weight1", "Wght1Unit", "Weight2", "Wght2Unit", "Factor1", "Factor2", "Factor3", "Factor4",
    "BaseType", "BaseEntry", "BaseLine", "Volume", "VolUnit", "Width1", "Wdth1Unit", "Width2", "Wdth2Unit", "Address",
    "TaxCode", "TaxType", "TaxStatus", "BackOrdr", "FreeTxt", "TrnsCode", "CEECFlag", "ToStock", "ToDiff", "WtLiable",
    "DeferrTax", "unitMsr", "NumPerMsr", "LineTotal", "VatPrcnt", "VatSum", "ConsumeFCT", "ExciseAmt", "CountryOrg", "SWW",
    "TranType", "DistribExp", "TotalFrgn", "CFOPCode", "CSTCode", "Usage", "TaxOnly", "PriceBefDi", "LineStatus", "PackQty",
    "LineType", "CogsOcrCod", "CogsAcct", "ChgAsmBoMW", "GrossBuyPr", "GrossBase", "GPTtlBasPr", "OcrCode2", "OcrCode3", "OcrCode4",
    "OcrCode5", "Text", "LocCode", "ActDelDate", "ExLineNo", "PQTReqDate", "PQTReqQty", "CogsOcrCo2", "CogsOcrCo3", "CogsOcrCo4",
    "CogsOcrCo5", "CSTfIPI", "CSTfPIS", "CSTfCOFINS", "CredOrigin", "NoInvtryMv", "AgrNo", "AgrLnNum", "ActBaseEnt", "ActBaseLn",
    "DocEntry", "Surpluses", "DefBreak", "Shortages", "NeedQty", "PartRetire", "RetireQty", "RetireAPC", "ThirdParty", "PoNum",
    "PoItmNum", "ExpType", "ExpUUID", "ExpOpType", "LicTradNum", "GrssProfit", "GrssProfFC", "GrssProfSC", "SpecPrice", "UomEntry",
    "InvQty", "PrntLnNum", "Incoterms", "TransMod", "NatOfTrans", "ISDtCryImp", "ISDtRgnImp", "ISOrCryExp", "ISOrRgnExp", "InvQtyOnly",
    "FreeChrgBP", "SacEntry", "HsnEntry", "GPBefDisc", "GTotal", "GTotalFC", "NCMCode", "NVECode", "IndEscala", "CtrSealQty",
    "CNJPMan", "CESTCode", "UFFiscBene", "RevCharge", "ShipToCode", "ShipToDesc", "OwnerCode", "ExtTaxRate", "ExtTaxSum", "ExtTaxSumF",
    "ExtTaxSumS", "StdItemId", "CommClass", "UnencReasn", "CUSplit", "U_ZL_SlpCode", "U_ZL_SlpName", "U_ZL_CardCode", "U_ZL_CardName", "U_ZL_InvoiceNum",
    "U_ZL_ItemCode", "U_ZL_ItemName", "U_BonusItem", "U_HasAlternativeItem", "U_PermitNum", "U_Has_EPoint", "U_PI_Qty", "U_PI_Price", "U_Redemption", "U_EPoint_Redeemed",
    "U_Rebate_RefNum", "U_SLS_Remarks1", "U_SLS_Remarks2", "U_SLS_Remarks3", "U_SLS_Remarks4", "U_SLS_Remarks5", "U_GSTSGDRate", "U_GSTSGDAmt", "U_Related_SO_Number", "U_Related_SO_LineNum",
    "U_QUT_Calc_BOM_Cost", "U_RDR_HasSpcTerms", "U_RDR_SpcPayTerms", "U_TRF_SerialNum", "U_ELTO_DefIncInd", "U_ELTO_ConStrDat", "U_ELTO_ConEndDat", "U_ELTO_FreqBill", "U_ELTO_FreqAdvRev", "U_ELTO_BillStrDat",
    "U_ELTO_LseQty", "U_ELTO_CustPO", "U_ELTO_EstBillAmt", "U_ELTO_BillId", "U_ELTO_BillLine", "U_ELTO_NoCustPOApp", "U_ELTO_Lease", "U_ELTO_LseItmCde", "U_ELTO_LseItmDsc", "U_ELTO_LseWhs",
    "U_ELTO_LseDlvQty", "U_ELTO_LseRtnQty", "U_ELTO_RevSchType", "U_ELTO_ContractQty", "U_ELTO_QtyConPrice", "U_ELTO_Prorata", "U_ELTO_BillPeriodFr", "U_ELTO_BillPeriodTo", "U_ELTO_BaseDocEntry", "U_ELTO_BaseDocNum",
    "U_ELTO_BaseRow", "U_FIN_Shown", "U_PL_CartonNum", "U_PL_KgCarton", "U_PL_MYRegNum", "U_TRF_VIT_Prc", "U_TRF_SellingPrice", "U_UnitPriceBP", "U_POCountry", "U_Invoice_No",
    "U_Invoice_Item", "U_Return_Ref_No", "U_P_Customer_Code", "U_ZP_Item_Code", "U_Credit_Reason", "U_List_Price", "U_ZP_Customer_Code", "U_Customer_Name", "U_ZP_Invoice_Item", "U_Customer_State",
    "U_Cust_Territory", "U_TempDivision", "U_RecLines", "U_COO"
]


# ============================================================================
# VALIDATION FUNCTIONS
# ============================================================================
def validate_date_format(date_val) -> bool:
    """Check if date is in YYYYMMDD format"""
    if pd.isna(date_val):
        return False
    date_str = str(int(date_val)) if isinstance(date_val, float) else str(date_val)
    if len(date_str) != 8:
        return False
    if not date_str.isdigit():
        return False
    try:
        y, m, d = int(date_str[:4]), int(date_str[4:6]), int(date_str[6:8])
        if y < 1900 or y > 2100 or m < 1 or m > 12 or d < 1 or d > 31:
            return False
        return True
    except:
        return False


def validate_orders(df_orders: pd.DataFrame) -> List[str]:
    """Validate order headers and return list of errors"""
    errors = []
    
    if df_orders.empty:
        errors.append("No orders found in 'Sales Order Entry' sheet")
        return errors
    
    seen_doc_nums = set()
    
    for idx, row in df_orders.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        
        # Skip empty rows
        if pd.isna(row.get('Order #')) or str(row.get('Order #')).strip() == '':
            continue
        
        doc_num = str(row.get('Order #')).strip()
        
        # Check duplicate
        if doc_num in seen_doc_nums:
            errors.append(f"Row {row_num}: Duplicate Order # '{doc_num}'")
        seen_doc_nums.add(doc_num)
        
        # Required fields
        if pd.isna(row.get('Document Date')) or str(row.get('Document Date')).strip() == '':
            errors.append(f"Row {row_num}: Document Date is required")
        elif not validate_date_format(row.get('Document Date')):
            errors.append(f"Row {row_num}: Document Date must be YYYYMMDD format")
        
        if pd.isna(row.get('Due Date')) or str(row.get('Due Date')).strip() == '':
            errors.append(f"Row {row_num}: Due Date is required")
        elif not validate_date_format(row.get('Due Date')):
            errors.append(f"Row {row_num}: Due Date must be YYYYMMDD format")
        
        if pd.isna(row.get('Customer Code')) or str(row.get('Customer Code')).strip() == '':
            errors.append(f"Row {row_num}: Customer Code is required")
        
        if pd.isna(row.get('Sales Code')) or str(row.get('Sales Code')).strip() == '':
            errors.append(f"Row {row_num}: Sales Code is required")
    
    return errors


def validate_lines(df_lines: pd.DataFrame, valid_order_nums: set) -> List[str]:
    """Validate line items and return list of errors"""
    errors = []
    
    if df_lines.empty:
        errors.append("No line items found in 'Line Items Entry' sheet")
        return errors
    
    for idx, row in df_lines.iterrows():
        row_num = idx + 2
        
        # Skip empty rows
        if pd.isna(row.get('Parent Order #')) or str(row.get('Parent Order #')).strip() == '':
            continue
        
        parent_key = str(row.get('Parent Order #')).strip()
        
        # Check parent exists
        if parent_key not in valid_order_nums:
            errors.append(f"Row {row_num}: Parent Order # '{parent_key}' not found in Order Headers")
        
        # Required fields
        if pd.isna(row.get('Line #')):
            errors.append(f"Row {row_num}: Line # is required")
        
        if pd.isna(row.get('Item Code')) or str(row.get('Item Code')).strip() == '':
            errors.append(f"Row {row_num}: Item Code is required")
        
        if pd.isna(row.get('Quantity')):
            errors.append(f"Row {row_num}: Quantity is required")
        elif float(row.get('Quantity', 0)) <= 0:
            errors.append(f"Row {row_num}: Quantity must be positive")
        
        if pd.isna(row.get('Warehouse')) or str(row.get('Warehouse')).strip() == '':
            errors.append(f"Row {row_num}: Warehouse is required")
        
        if pd.isna(row.get('Sales Code')) or str(row.get('Sales Code')).strip() == '':
            errors.append(f"Row {row_num}: Sales Code is required")
        
        if pd.isna(row.get('Account Code')) or str(row.get('Account Code')).strip() == '':
            errors.append(f"Row {row_num}: Account Code is required")
        
        if pd.isna(row.get('VAT Group')) or str(row.get('VAT Group')).strip() == '':
            errors.append(f"Row {row_num}: VAT Group is required")
    
    return errors


# ============================================================================
# FILE GENERATION FUNCTIONS
# ============================================================================
def format_date(val) -> str:
    """Format date value to YYYYMMDD string"""
    if pd.isna(val):
        return ""
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val).strip()


def safe_str(val) -> str:
    """Safely convert value to string"""
    if pd.isna(val):
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def generate_ordr_file(df_orders: pd.DataFrame) -> str:
    """Generate ORDR.txt content"""
    lines = []
    
    # Header row 1
    lines.append('\t'.join(ORDR_HEADER_ROW1))
    
    # Header row 2
    lines.append('\t'.join(ORDR_HEADER_ROW2))
    
    # Data rows
    for _, row in df_orders.iterrows():
        if pd.isna(row.get('Order #')) or str(row.get('Order #')).strip() == '':
            continue
        
        # Initialize all columns to empty
        data = [''] * ORDR_TOTAL_COLS
        
        # Populate fields
        data[ORDR_FIELDS['DocNum']] = safe_str(row.get('Order #'))
        data[ORDR_FIELDS['DocType']] = 'dDocument_Items'
        data[ORDR_FIELDS['DocDate']] = format_date(row.get('Document Date'))
        data[ORDR_FIELDS['DocDueDate']] = format_date(row.get('Due Date'))
        data[ORDR_FIELDS['CardCode']] = safe_str(row.get('Customer Code'))
        data[ORDR_FIELDS['SalesPersonCode']] = safe_str(row.get('Sales Code'))
        data[ORDR_FIELDS['U_AllowDOInv_Dep']] = 'Y'
        data[ORDR_FIELDS['U_GSTSGDRate']] = safe_str(row.get('Branch ID', 9))
        
        lines.append('\t'.join(data))
    
    return '\r\n'.join(lines) + '\r\n'


def generate_rdr1_file(df_lines: pd.DataFrame) -> str:
    """Generate RDR1.txt content"""
    lines = []
    
    # Header row 1
    lines.append('\t'.join(RDR1_HEADER_ROW1))
    
    # Header row 2
    lines.append('\t'.join(RDR1_HEADER_ROW2))
    
    # Data rows
    for _, row in df_lines.iterrows():
        if pd.isna(row.get('Parent Order #')) or str(row.get('Parent Order #')).strip() == '':
            continue
        
        # Initialize all columns to empty
        data = [''] * RDR1_TOTAL_COLS
        
        # Populate fields
        data[RDR1_FIELDS['ParentKey']] = safe_str(row.get('Parent Order #'))
        data[RDR1_FIELDS['LineNum']] = safe_str(row.get('Line #'))
        data[RDR1_FIELDS['ItemCode']] = safe_str(row.get('Item Code'))
        data[RDR1_FIELDS['Quantity']] = safe_str(row.get('Quantity'))
        data[RDR1_FIELDS['Price']] = safe_str(row.get('Unit Price', ''))
        data[RDR1_FIELDS['WarehouseCode']] = safe_str(row.get('Warehouse'))
        data[RDR1_FIELDS['SalesPersonCode']] = safe_str(row.get('Sales Code'))
        data[RDR1_FIELDS['AccountCode']] = safe_str(row.get('Account Code'))
        data[RDR1_FIELDS['VatGroup']] = safe_str(row.get('VAT Group'))
        
        # Dimension codes
        dim1 = safe_str(row.get('Dim 1', ''))
        dim2 = safe_str(row.get('Dim 2', ''))
        dim3 = safe_str(row.get('Dim 3', ''))
        dim4 = safe_str(row.get('Dim 4', ''))
        dim5 = safe_str(row.get('Dim 5', ''))
        
        data[RDR1_FIELDS['CostingCode']] = dim1
        data[RDR1_FIELDS['CostingCode2']] = dim2
        data[RDR1_FIELDS['CostingCode3']] = dim3
        data[RDR1_FIELDS['CostingCode4']] = dim4
        data[RDR1_FIELDS['CostingCode5']] = dim5
        
        # COGS dimension codes (mirror)
        data[RDR1_FIELDS['COGSCostingCode']] = dim1
        data[RDR1_FIELDS['COGSCostingCode2']] = dim2
        data[RDR1_FIELDS['COGSCostingCode3']] = dim3
        data[RDR1_FIELDS['COGSCostingCode4']] = dim4
        data[RDR1_FIELDS['COGSCostingCode5']] = dim5
        
        # Optional fields
        data[RDR1_FIELDS['U_PermitNum']] = safe_str(row.get('Permit #', ''))
        data[RDR1_FIELDS['U_GSTSGDRate']] = safe_str(row.get('Branch', 9))
        
        lines.append('\t'.join(data))
    
    return '\r\n'.join(lines) + '\r\n'


# ============================================================================
# TEMPLATE GENERATION (embedded - no external file needed)
# ============================================================================
# Sales Employees data (embedded for standalone deployment)
SALES_EMPLOYEES = [
    (260, "CORP - Clare Chan"), (149, "CORP - Hanh"), (2, "CORP - Jimmy Ang"),
    (146, "CORP - Kee Cheng"), (147, "CORP - Moe"), (214, "CORP - Neoasia Connect"),
    (148, "CORP - Nga"), (48, "MED I - Adeline Chia"), (198, "MED I - Alicia Chua"),
    (237, "MED I - Andrea Seet"), (159, "MED I - Buying out for listing"),
    (181, "MED I - Cam Van"), (225, "MED I - Cao Tuyet Hien"), (193, "MED I - Carmen Teng"),
    (24, "MED I - Charissa Ngai"), (156, "MED I - Dieu An"), (230, "MED I - Duong Tu Nhi"),
    (103, "MED I - Eileen Chua"), (224, "MED I - Gi Wen"), (6, "MED I - Glynise Peh"),
    (75, "MED I - Heng Hui"), (211, "MED I - Jamie Chow"), (202, "MED I - Jenny Chu Chin Lee (MAL)"),
    (128, "MED I - Joceline Siaw"), (196, "MED I - Kher Yiow"), (258, "MED I - Le Thi Cam Tuyen"),
    (163, "MED I - Mai Huong"), (155, "MED I - MISC"), (262, "MED I - Nguyen Hoang Dieu Linh"),
    (256, "MED I - Nguyen Kim Dung"), (257, "MED I - Nguyen Thi Nguyet Anh"),
    (184, "MED I - Pham Thi Thuy Duong"), (143, "MED I - Sandy Li"), (190, "MED I - Sharry Woon"),
    (123, "MED I - Thi Hoan"), (106, "MED I - Thi Thanh"), (250, "MED I - Thina Som"),
    (121, "MED I - Thu Thao"), (261, "MED I - Tran Thi Ngoc Ha"), (255, "MED I - Tran Thi Thu Uyen"),
    (223, "MED I - Valerie Tan"), (240, "MED I - Vu Thi Thanh Huyen"), (158, "MED I - Withdraw"),
    (212, "MED II - Adeline Chia"), (254, "MED II - Cao Thi Huyen Anh"),
    (25, "MED II - Charissa Ngai"), (265, "MED II - Dang Trang Nhung"),
    (232, "MED II - Duong Tu Nhi"), (204, "MED II - Eric Chua"), (7, "MED II - Glynise Peh"),
    (80, "MED II - Heng Hui"), (264, "MED II - Le Thi Cam Tuyen"), (161, "MED II - Maggie Chua"),
    (234, "MED II - Mai Huong"), (235, "MED II - Melissa Lai"),
    (263, "MED II - Nguyen Hoang Dieu Linh"), (244, "MED II - Phealey"),
    (245, "MED II - Thina Som"), (226, "MED II - Tran Thi Minh Ngoc"),
    (259, "MED II - Tran Thi Ngoc Huyen"), (177, "MED II - Tu Em"), (227, "MED II - Valerine Siew"),
    (18, "MED III - Cliff Ang"), (114, "MED III - Daphne YC"), (101, "MED III - Glynise Peh"),
    (105, "MED III - Heng Hui"), (192, "MED III - Kef Lim"), (206, "MED III - Maggie Chua"),
    (210, "MED III - Robylyn Agustin"), (23, "MED III - Ryan Seng"), (233, "MED III - Stacy Toh"),
    (253, "MED III - Thina Som"), (152, "MED III - Zhi Yang"), (215, "OMNI - Bernice Tiew"),
    (167, "OMNI - Goh Kah Mun"), (199, "OMNI - Joey Tan"), (207, "OMNI - Tan Si Jie"),
]

@st.cache_data
def generate_template() -> bytes:
    """Generate Excel template in memory"""
    # Styles
    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
    INPUT_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    wb = Workbook()
    
    # Sheet 1: Sales Order Entry
    ws_orders = wb.active
    ws_orders.title = "Sales Order Entry"
    
    order_headers = ['Order #', 'Document Date', 'Due Date', 'Customer Code', 'Sales Code', 'Branch ID']
    for col, header in enumerate(order_headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
    
    ws_orders.column_dimensions['A'].width = 12
    ws_orders.column_dimensions['B'].width = 15
    ws_orders.column_dimensions['C'].width = 15
    ws_orders.column_dimensions['D'].width = 15
    ws_orders.column_dimensions['E'].width = 12
    ws_orders.column_dimensions['F'].width = 12
    
    for row in range(2, 102):
        for col in range(1, 7):
            cell = ws_orders.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        ws_orders.cell(row=row, column=6, value=9)
    
    ws_orders.freeze_panes = 'A2'
    
    # Sheet 2: Line Items Entry
    ws_lines = wb.create_sheet("Line Items Entry")
    
    line_headers = ['Parent Order #', 'Line #', 'Item Code', 'Quantity', 'Unit Price',
                    'Warehouse', 'Sales Code', 'Account Code', 'VAT Group',
                    'Dim 1', 'Dim 2', 'Dim 3', 'Dim 4', 'Dim 5', 'Permit #', 'Branch']
    
    for col, header in enumerate(line_headers, 1):
        cell = ws_lines.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    widths = [14, 8, 15, 10, 12, 10, 12, 12, 10, 8, 8, 8, 10, 8, 15, 8]
    for i, w in enumerate(widths, 1):
        ws_lines.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = w
    
    for row in range(2, 202):
        for col in range(1, 17):
            cell = ws_lines.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
        ws_lines.cell(row=row, column=16, value=9)
    
    # Dropdowns
    dv_whs = DataValidation(type='list', formula1='"SG01,SG02,MY01,MY02,VN01,ID01,TH01,PH01"', allow_blank=True)
    ws_lines.add_data_validation(dv_whs)
    dv_whs.add('F2:F201')
    
    dv_vat = DataValidation(type='list', formula1='"SR,ZR,ES,OS,TX,IM,DS"', allow_blank=True)
    ws_lines.add_data_validation(dv_vat)
    dv_vat.add('I2:I201')
    
    ws_lines.freeze_panes = 'A2'
    
    # Sheet 3: Sales Employees Reference
    ws_emp = wb.create_sheet("SalesEmployees")
    
    ws_emp['A1'] = 'Code'
    ws_emp['B1'] = 'Sales Employee Name'
    ws_emp['A1'].font = HEADER_FONT
    ws_emp['B1'].font = HEADER_FONT
    ws_emp['A1'].fill = HEADER_FILL
    ws_emp['B1'].fill = HEADER_FILL
    
    ws_emp.column_dimensions['A'].width = 10
    ws_emp.column_dimensions['B'].width = 40
    
    for i, (code, name) in enumerate(SALES_EMPLOYEES, 2):
        ws_emp.cell(row=i, column=1, value=code)
        ws_emp.cell(row=i, column=2, value=name)
    
    # Sheet 4: Instructions
    ws_inst = wb.create_sheet("Instructions")
    ws_inst.column_dimensions['A'].width = 80
    
    instructions = [
        "NeoAsia Sales Order DTW Template",
        "=" * 50,
        "",
        "HOW TO USE:",
        "1. Fill 'Sales Order Entry': One row per order, YYYYMMDD dates",
        "2. Fill 'Line Items Entry': Parent Order # must match Order #",
        "3. Look up Sales Codes in 'SalesEmployees' sheet",
        "4. Upload to DTW Web App and download generated files",
        "5. Import ORDR.txt and RDR1.txt into SAP DTW",
        "",
        "NOTES:",
        "- Green cells = data entry areas",
        "- Line # starts at 1 for each order",
        "- Use Warehouse and VAT dropdowns in Line Items",
    ]
    
    for i, line in enumerate(instructions, 1):
        ws_inst.cell(row=i, column=1, value=line)
    
    wb.move_sheet("Instructions", offset=-3)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ============================================================================
# STREAMLIT UI
# ============================================================================
def main():
    # Header
    st.title("📦 NeoAsia Sales Order DTW Tool")
    st.markdown("**Generate SAP DTW import files from Excel data**")
    
    st.divider()
    
    # Instructions
    with st.expander("📖 How to Use", expanded=False):
        st.markdown("""
        ### Step-by-Step Guide
        
        1. **Download the Excel template** (link below)
        2. **Fill in your data:**
           - `Sales Order Entry` sheet: One row per order
           - `Line Items Entry` sheet: Line items for each order
        3. **Upload the filled Excel file** using the uploader below
        4. **Review validation results** - fix any errors shown
        5. **Click Generate** to create the DTW files
        6. **Download the ZIP** containing ORDR.txt and RDR1.txt
        7. **Copy files to Terminal Server** and import via SAP DTW
        
        ### Important Rules
        - Date format must be **YYYYMMDD** (e.g., 20260115)
        - Each Order # must be **unique**
        - Parent Order # in Line Items must **match** an Order # in Sales Order Entry
        - Line numbers start at **1** for each order
        """)
    
    # Template download
    st.markdown("### 📥 Download Template")
    st.info("First, download and fill in the Excel template, then upload it below.")
    
    # Generate template download
    template_buffer = generate_template()
    st.download_button(
        label="📄 Download Excel Template",
        data=template_buffer,
        file_name="DTW_Sales_Order_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # File upload
    st.markdown("### 📤 Upload Filled Template")
    uploaded_file = st.file_uploader(
        "Choose your filled Excel file",
        type=['xlsx', 'xls'],
        help="Upload the Excel file with your Sales Orders and Line Items"
    )
    
    if uploaded_file is not None:
        try:
            # Read Excel file
            with st.spinner("Reading Excel file..."):
                xl = pd.ExcelFile(uploaded_file)
                
                # Check for required sheets
                required_sheets = ['Sales Order Entry', 'Line Items Entry']
                missing_sheets = [s for s in required_sheets if s not in xl.sheet_names]
                
                if missing_sheets:
                    st.error(f"❌ Missing required sheets: {', '.join(missing_sheets)}")
                    st.info("Your Excel file must have sheets named 'Sales Order Entry' and 'Line Items Entry'")
                    return
                
                df_orders = pd.read_excel(xl, sheet_name='Sales Order Entry')
                df_lines = pd.read_excel(xl, sheet_name='Line Items Entry')
            
            # Display data preview
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### Orders Preview")
                # Count non-empty orders
                order_count = df_orders[df_orders['Order #'].notna()].shape[0]
                st.metric("Orders Found", order_count)
                if order_count > 0:
                    st.dataframe(df_orders[df_orders['Order #'].notna()].head(10), use_container_width=True)
            
            with col2:
                st.markdown("#### Line Items Preview")
                # Count non-empty lines
                line_count = df_lines[df_lines['Parent Order #'].notna()].shape[0]
                st.metric("Line Items Found", line_count)
                if line_count > 0:
                    st.dataframe(df_lines[df_lines['Parent Order #'].notna()].head(10), use_container_width=True)
            
            st.divider()
            
            # Validation
            st.markdown("### ✅ Validation")
            
            with st.spinner("Validating data..."):
                # Get valid order numbers
                valid_orders = set()
                for _, row in df_orders.iterrows():
                    if pd.notna(row.get('Order #')):
                        valid_orders.add(str(row.get('Order #')).strip())
                
                order_errors = validate_orders(df_orders)
                line_errors = validate_lines(df_lines, valid_orders)
                
                all_errors = order_errors + line_errors
            
            if all_errors:
                st.error(f"❌ Found {len(all_errors)} validation error(s)")
                for error in all_errors[:20]:  # Show first 20
                    st.markdown(f"- {error}")
                if len(all_errors) > 20:
                    st.markdown(f"*... and {len(all_errors) - 20} more errors*")
                st.warning("Please fix the errors above and re-upload the file.")
            else:
                st.success("✅ All validation checks passed!")
                
                st.divider()
                
                # Generate button
                st.markdown("### 📁 Generate DTW Files")
                
                if st.button("🚀 Generate ORDR.txt and RDR1.txt", type="primary", use_container_width=True):
                    with st.spinner("Generating files..."):
                        # Generate content
                        ordr_content = generate_ordr_file(df_orders)
                        rdr1_content = generate_rdr1_file(df_lines)
                        
                        # Create timestamp
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        
                        # Create ZIP in memory
                        zip_buffer = io.BytesIO()
                        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                            zf.writestr(f'ORDR_{timestamp}.txt', ordr_content)
                            zf.writestr(f'RDR1_{timestamp}.txt', rdr1_content)
                        
                        zip_buffer.seek(0)
                    
                    st.success("✅ Files generated successfully!")
                    
                    # Download button
                    st.download_button(
                        label="📥 Download DTW Files (ZIP)",
                        data=zip_buffer,
                        file_name=f"DTW_Sales_Orders_{timestamp}.zip",
                        mime="application/zip",
                        use_container_width=True
                    )
                    
                    # Summary
                    st.info(f"""
                    **Generated Files:**
                    - `ORDR_{timestamp}.txt` - {order_count} order(s)
                    - `RDR1_{timestamp}.txt` - {line_count} line item(s)
                    
                    **Next Steps:**
                    1. Extract the ZIP file
                    2. Copy both .txt files to the Terminal Server
                    3. Open SAP DTW > Import > Transactional Data > Sales Orders
                    4. Select both files and run
                    """)
        
        except Exception as e:
            st.error(f"❌ Error reading file: {str(e)}")
            st.info("Please ensure your file is a valid Excel file with the correct sheet names.")
    
    # Footer
    st.divider()
    st.markdown(
        "<div style='text-align: center; color: #888;'>NeoAsia Sales Order DTW Tool v1.0.0 | Business Analytics Team</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
